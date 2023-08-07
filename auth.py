# Import required modules and classes
import pandas as pd
from openpyxl.drawing.image import Image
from matplotlib import pyplot as plt
from openpyxl.workbook import Workbook
from rapidfuzz import fuzz as rapidfuzz
from rapidfuzz import process as rapidfuzz_process
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, jsonify
from .models import User
from . import db
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import login_user, login_required, logout_user, current_user
import os
import swifter

# Read the admin key from 'static/admin_key.txt'
with open('static/admin_key.txt', 'r') as f:
    adminKeyValid = f.read()

# Create the 'auth' Blueprint
auth = Blueprint('auth', __name__)


# Authentication - Login
@auth.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Get the email and password from the login form submitted via POST request
        email = request.form.get('email')
        password = request.form.get('password')

        # Query the database for the user with the provided email
        user = User.query.filter_by(email=email).first()

        # Check if the user exists
        if user:
            # If the user exists, check if the provided password matches the hashed password in the database
            if check_password_hash(user.password, password):
                # If the passwords match, display a success message, log in the user, and redirect to the home page
                flash('Logged in successfully!', category='success')
                login_user(user, remember=True)
                return redirect(url_for('views.home'))
            else:
                # If the passwords don't match, display an error message
                flash('Incorrect password, try again.', category='error')
        else:
            # If the user doesn't exist, display an error message
            flash('Email does not exist.', category='error')

    # If the request method is GET or there was an error in login (e.g., incorrect password), render the login page
    return render_template('login.html', user=current_user)


# Logout - Route for logging out the user
@auth.route('/logout')
@login_required  # Requires the user to be logged in to access this view
def logout():
    # Use Flask-Login's 'logout_user()' function to log out the current user
    logout_user()

    # After logout, redirect the user to the login page
    return redirect(url_for('auth.login'))


# Sign-up - Route for user registration
@auth.route('/sign-up', methods=['GET', 'POST'])
def sign_up():
    if request.method == 'POST':
        # Get user inputs from the sign-up form submitted via POST request
        email = request.form.get('email')
        password1 = request.form.get('password1')
        password2 = request.form.get('password2')
        adminKey = request.form.get('adminKey')

        # Query the database for a user with the provided email
        user = User.query.filter_by(email=email).first()

        # Check if the email is already in use by an existing user
        if user:
            flash('Email already exists.', category='error')
        # Validate the email length
        elif len(email) < 4:
            flash('Email must be greater than 3 characters.', category='error')
        # Check if the two provided passwords match
        elif password1 != password2:
            flash('Passwords don\'t match.', category='error')
        else:
            # Check if the provided admin key matches the valid admin key
            if adminKey != adminKeyValid and adminKey != '':
                flash('Admin key is invalid.', category='error')
            else:
                # If the admin key is valid, set the 'isAdmin' flag to True
                isAdmin = adminKey == adminKeyValid

                # Create a new User object with the provided email and hashed password
                new_user = User(email=email, password=generate_password_hash(password1, method='sha256'),
                                isAdmin=isAdmin)

                # Add the new user to the database and commit the changes
                db.session.add(new_user)
                db.session.commit()

                # Log in the new user using Flask-Login's 'login_user' function
                login_user(new_user, remember=True)

                # Show a success message and redirect the user to the home page
                flash('Account created!', category='success')
                return redirect(url_for('views.home'))

    # If the request method is GET or there was an error in the sign-up process, render the sign-up page
    return render_template('sign_up.html', user=current_user)


# Upload Records - Route for uploading and processing records
@auth.route('/uploadRecords', methods=['GET', 'POST'])
def upload_records():
    # Get the list of database files in the 'static/databases/' directory
    db_files = os.listdir('static/databases/')

    if request.method == 'POST':
        # Get the file and database name from the form submitted via POST request
        file = request.files['record_file']
        database_name = request.form.get('database_name')

        # Check if a file was selected for upload
        if file.filename == '':
            flash('No selected file', category='error')
            return redirect(url_for('views.uploadRecords'))

        # Check if the specified database name exists in the 'static/databases/' directory
        if database_name not in db_files:
            flash('Database name not found', category='error')
            return redirect(url_for('views.uploadRecords'))

        # If a file is uploaded and the database name is valid, proceed with file processing
        if file:
            file_path = 'static/records/' + current_user.email + '_' + file.filename
            database_path = 'static/databases/' + database_name
            file.save(file_path)

            try:
                df1 = pd.read_excel(file_path, nrows=100)
            except Exception as e:
                flash(f'Cannot read file: {e}', category='error')
                return redirect(url_for('views.uploadRecords'))

            # Check if the file has only one column
            if len(df1.columns) != 1:
                flash('File must have only one column.', category='error')
                return redirect(url_for('views.uploadRecords'))
            else:
                flash('File uploaded successfully!', category='success')

            # Load the database file into a DataFrame 'df2'
            df2 = pd.read_excel(database_path)

            # Convert the database names to lowercase for case-insensitive matching
            df2[df2.columns[0]] = df2[df2.columns[0]].str.lower()

            # Create a DataFrame 'result_df' to store the results of the fuzzy string matching
            result_df = pd.DataFrame()
            result_df['name'] = df1[df1.columns[0]].str.lower()
            result_df.dropna(inplace=True)

            # Configure the fuzzy string matching settings
            score_method = rapidfuzz.ratio
            n_matches = 3

            # Define a function to find the closest matches for each name in 'result_df'
            def get_closest_match(curr_row: pd.Series) -> pd.Series:
                closest_matches = rapidfuzz_process.extract(curr_row['name'], df2[df2.columns[0]], scorer=score_method,
                                                            limit=n_matches)
                for i, match in enumerate(closest_matches):
                    curr_row[f'match{i + 1}'] = match[0]
                    curr_row[f'match_score{i + 1}'] = match[1]

                return curr_row

            # Apply the matching function to 'result_df' rows in parallel using 'swifter'
            result_df = result_df.swifter.apply(get_closest_match, axis=1)
            result_df = result_df.dropna()

            # Create a figure for the visualizations with a large size
            fig = plt.figure(figsize=(40, 50))

            # Adjust the vertical space between subplots
            plt.subplots_adjust(hspace=.5)

            # Set the font size for subplot titles
            plt.rcParams['axes.titlesize'] = 13

            # Set the background color of the figure to white
            fig.patch.set_facecolor('white')

            # Function to format the text for displaying statistics
            def get_text_str(mu, minimum, maximum):
                return '\n'.join((
                    r'$\mu=%.2f$' % (mu,),
                    r'min=%.2f' % (minimum,),
                    r'max=%.2f' % (maximum,)))

            # Function to create a histogram with a threshold
            def hist_threshold(hist_ax, column_name: str, title: str, greater=False):
                if greater:
                    hist_data = result_df[result_df[column_name] >= 90][column_name]
                else:
                    hist_data = result_df[result_df[column_name] < 90][column_name]

                mu = hist_data.mean()
                minimum = hist_data.min()
                maximum = hist_data.max()
                hist_ax.hist(hist_data, int(90 / 10))
                hist_ax.set_xlabel('Score')
                hist_ax.set_ylabel('Count')
                hist_ax.set_title(title)
                props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
                hist_ax.text(0.03, 0.9, get_text_str(mu, minimum, maximum), transform=hist_ax.transAxes, fontsize=11,
                             verticalalignment='top', bbox=props)

            # Function to create a pie chart with a threshold
            def pie_threshold(pie_ax, column_name: str, title: str, thresh: int):
                pie_data = result_df[column_name]
                pie_data = pie_data.apply(lambda x: 1 if x >= thresh else 0)
                try:
                    pie_ax.pie(pie_data.value_counts(), labels=[f'<{thresh}%', f'{thresh}%'], autopct='%1.1f%%')
                except:
                    pie_ax.pie(pie_data.value_counts(), labels=[f'{thresh}%'], autopct='%1.1f%%')
                pie_ax.set_title(title)

            # Define the number of columns and rows for the subplots grid
            n_cols = 3
            n_rows = 6

            # Define the threshold value for pie charts
            threshold = 95

            # Loop over the number of matches and create subplots with histograms and pie charts
            for n in range(1, n_matches + 1):
                ax = plt.subplot(n_rows, n_cols, (n - 1) * 3 + 1)
                hist_threshold(ax, f'match_score{n}', f'Distribution of scores below the threshold for match {n}')
                ax = plt.subplot(n_rows, n_cols, (n - 1) * 3 + 2)
                hist_threshold(ax, f'match_score{n}', f'Distribution of scores above the threshold for match {n}',
                               greater=True)
                ax = plt.subplot(n_rows, n_cols, (n - 1) * 3 + 3)
                if n == 1:
                    pie_threshold(ax, f'match_score{n}', f'Distribution of scores that are 100% for match {n}',
                                  threshold)
                else:
                    pie_threshold(ax, f'match_score{n}', f'Distribution of scores that are {threshold}% for match {n}',
                                  threshold)

            # Create subplots for displaying the average scores
            ax = plt.subplot(n_rows, n_cols, n_matches * 3 + 1)
            data = result_df[[f'match_score{n}' for n in range(1, n_matches + 1)]].mean()
            ax.bar(data.index, data.values)
            ax.set_xlabel('Match')
            ax.set_ylabel('Score')
            ax.set_title('Average score for each match')
            plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')

            ax = plt.subplot(n_rows, n_cols, n_matches * 3 + 2)
            data = result_df[[f'match_score{n}' for n in range(1, n_matches + 1)]].mean(axis=1)
            ax.plot(data.index, data.values)
            ax.set_xlabel('Insurer name')
            ax.set_ylabel('Score')
            ax.set_title('Average score for each insurer name')

            ax = plt.subplot(n_rows, n_cols, n_matches * 3 + 3)
            data = result_df[[f'match_score{n}' for n in range(1, n_matches + 1)]]
            ax.boxplot(data.values)
            ax.set_xlabel('Match #')
            ax.set_ylabel('Score')
            ax.set_title('Box and whisker chart for each match')

            # Set the file path for saving the generated visualization image
            img_path = f'static/matches/{current_user.email}_match.png'

            # Save the figure with all the subplots to the specified image path
            plt.savefig(img_path)

            # Create a new Excel workbook
            wb = Workbook()

            # Remove the default 'Sheet' created by the workbook
            wb.remove(wb['Sheet'])

            # Create a new worksheet for storing the data from the result DataFrame
            dataframe_sheet = wb.create_sheet('Results')

            # Create a new worksheet for storing the visualization image
            image_sheet = wb.create_sheet('Report')

            # Append the column headers of the result DataFrame to the 'Results' worksheet
            dataframe_sheet.append(result_df.columns.tolist())

            # Append each row of the result DataFrame to the 'Results' worksheet
            for row in result_df.values.tolist():
                dataframe_sheet.append(row)

            # Adjust the column widths in the 'Results' worksheet to fit the data
            for column_cells in dataframe_sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                dataframe_sheet.column_dimensions[column_cells[0].column_letter].width = length

            # Make the first row (header) bold in the 'Results' worksheet
            for cell in dataframe_sheet['1:1']:
                cell.font = cell.font.copy(bold=True)

            # Load the previously saved image into an Image object
            img = Image(img_path)

            # Set the width and height of the image in the Excel sheet
            img.width = 2000
            img.height = 2500

            # Add the image to the 'Report' worksheet at cell A1
            image_sheet.add_image(img, 'A1')

            # Set the download path for the Excel file
            download_path = f'static/matches/{current_user.email}_match.xlsx'

            # Save the workbook to the download path
            wb.save(download_path)

            # Display a success flash message to the user
            flash('File processed successfully!', category='success')

            # Render the 'uploadRecords.html' template with the updated variables
            return render_template('uploadRecords.html', user=current_user, database_files=db_files,
                                   download_ready=True, download_path=download_path)


# Define a new Flask route for downloading files
@auth.route('/download_file')
@login_required
def download_file():
    # Get the file path from the request arguments sent by the client
    file_path = request.args.get('file_path')
    # Use the Flask send_file function to send the file as an attachment for download
    return send_file(file_path, as_attachment=True)


# Define a new Flask route for the admin settings page
@auth.route('/admin', methods=['GET', 'POST'])
@login_required
def admin_settings():
    # Get the list of database files in the 'static/databases/' directory
    db_files = os.listdir('static/databases/')
    if request.method == 'POST':
        if 'database_file' in request.files and 'database_name' in request.form:
            # Get the uploaded file and form data from the request
            file = request.files['database_file']
            database_name = request.form.get('database_name')
            update_database = request.form.get('update_database')

            # Check if the database name already exists and user doesn't want to update
            if database_name in db_files and update_database != 'on':
                flash('Database name already exists', category='error')
                return redirect(url_for('views.admin'))

            # Check if the database name doesn't exist and user wants to update
            elif database_name not in db_files and update_database == 'on':
                flash('Database name not found', category='error')
                return redirect(url_for('views.admin'))

            # If a file was uploaded, save it to the specified database path
            if file:
                database_path = 'static/databases/' + database_name
                file.save(database_path)
                flash('Database uploaded successfully!', category='success')

        elif 'delete_database' in request.form:
            database_to_delete = request.form.get('delete_database')
            if database_to_delete:
                # Remove the database file from the 'static/databases/' directory
                os.remove(f'static/databases/{database_to_delete}')
                flash('Database deleted successfully!', category='success')

    # Render the 'admin.html' template with the current user data
    return render_template('admin.html', user=current_user, database_files=db_files)


# Define a new Flask route for the Single Match Analysis Page
@auth.route('/singleMatchAnalysis', methods=['GET', 'POST'])
@login_required
def single_match_analysis():
    # Get the list of database files in the 'static/databases/' directory
    db_files = os.listdir('static/databases/')
    if request.method == 'POST':
        # Get the input record (record_name)
        record_name = request.form.get('record_name')
        # Get the drop-down menu selection
        database_name = request.form.get('dataset_selection')
        # Get the number of matches
        num_matches = int(request.form.get('num_matches'))

        if database_name == 'All':
            database_files = os.listdir('static/databases/')
            first_column_values = [pd.read_excel(os.path.join('static/databases/', file_name)).iloc[:, 0].tolist() for
                                   file_name in database_files]
            df = pd.DataFrame(
                {"names": [value for sublist in first_column_values for value in sublist]})

        else:
            # Read the selected database file into a DataFrame
            df = pd.read_excel(f'static/databases/{database_name}')

        df = df.dropna()
        df['names'] = df['names'].astype(str)

        # Get the closest match using rapidfuzz
        closest_matches = rapidfuzz_process.extract(record_name, df['names'], limit=num_matches, scorer=rapidfuzz.ratio)
        closest_matches = [f'{match[0]}: {round(match[1], 2)}' for match in closest_matches]
        display_closest_matches = True

    return render_template('singleMatchAnalysis.html', user=current_user, database_files=db_files,
                           display_closest_matches=display_closest_matches, record_name=record_name,
                           closest_matches=closest_matches)


# Define a new Flask route for the User Close event
# This route is used to delete the user's files from the 'static/matches/' and 'static/records/' directory when they exit the site
@auth.route('/log_user_close', methods=['POST'])
def log_user_close():
    # Get the user_id from the request data (assuming it's sent as JSON)
    data = request.get_json()
    user_id = data.get('user_id')

    # Delete the user's files from the 'static/matches/' and 'static/records/' directory
    email = User.query.filter_by(id=user_id).first().email
    for file in os.listdir('static/matches/'):
        if email in file:
            os.remove(f'static/matches/{file}')
    for file in os.listdir('static/records/'):
        if email in file:
            os.remove(f'static/records/{file}')

    # Respond with a JSON response if needed
    return jsonify({'message': 'User close event logged successfully'})
